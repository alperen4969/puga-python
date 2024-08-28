import random
import numpy as np
import reshape_variables, repair_operator
import pandas as pd
import openpyxl
import copy
import var_limit
import scipy.io as sio


def export(x, y, z, h):
    x = np.vstack((x, y, z, h))
    wb = openpyxl.load_workbook("crossover_once_sonra.xlsx")
    sheet = wb["Sheet1"]
    # print(sheet.max_row)
    df = pd.DataFrame(x)
    with pd.ExcelWriter("crossover_once_sonra.xlsx", mode="a", engine="openpyxl", if_sheet_exists='overlay') as writer:
        df.to_excel(writer, index=False, header=False, sheet_name="Sheet1", startrow=sheet.max_row)


ind = 0


def crossover_op_pui(pop, opt, state):
    ub_eps = sio.loadmat("ub_eps.mat")["ub"]
    lb_eps = sio.loadmat("lb_eps.mat")["lb"]

    ctr = 0
    # fraction = opt['crossover_fraction']
    fraction = 2.0 / 12  # 0.166666666666667
    # range(0, len(pop), 2)
    for ind in range(0, len(pop), 2):  # pop_size should be even number
        # if ind == 22 :
        #     print("hey")
        child1, child2 = crossover_intermediate_pui(pop[ind], pop[ind + 1], fraction, opt)
        # child1['var'] = child1['var'].round()  # TODO: bunu aktif et
        # child2['var'] = child2['var'].round()
        child1['var'] = var_limit.var_limit_pui(child1['var'], lb_eps, ub_eps)
        child2['var'] = var_limit.var_limit_pui(child2['var'], lb_eps, ub_eps)

        pop[ind] = child1
        pop[ind + 1] = child2

    return pop


def crossover_intermediate_pui(parent1, parent2, fraction, opt):
    ratio = 1.2000  # matlab'ta baya kapsamlı gibi TODO

    # child1 = parent1.copy()
    # child2 = parent2.copy()
    child1 = copy.deepcopy(parent1)
    child2 = copy.deepcopy(parent2)
    # n_var = len(child1['var'][0])
    n_var = child1["var"].size
    ratio = np.array([ratio for _ in range(n_var)])
    ratio = np.reshape(ratio, (4, 3), order="F")
    # crs_flags = np.array([int(random.random() > fraction) for _ in range(n_var)])
    crs_flags = np.array([1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
    # rand_num = np.array([random.random() for _ in range(n_var)])
    rand_num = np.array([0.5 for _ in range(n_var)])
    crs_flags = np.reshape(crs_flags, (4, 3), order="F")
    rand_num = np.reshape(rand_num, (4, 3,), order="F")
    # [][0] idi...
    # child1['variables'][0] = [
    #     parent1['variables'][0][i] + crsflag[i] * randnum[i] * ratio * (parent2['variables'][0][i] - parent1['variables'][0][i])
    #     for i in range(nvar)]

    # crs_flags = np.array([1 for _ in range(n_var)])  # hepsi crossovera uğrar
    # test = parent1['variables'][0] + (crs_flags * rand_num * ratio * (parent2['variables'][0] - parent1['variables'][0]))

    test_0f = parent2['var'] - parent1['var']
    test0 = crs_flags * rand_num * ratio * (parent2['var'] - parent1['var'])
    child1['var'] = parent1['var'] + (crs_flags * rand_num * ratio * (parent2['var'] - parent1['var']))
    child2['var'] = parent2['var'] - (crs_flags * rand_num * ratio * (parent2['var'] - parent1['var']))

    # child2['variables'][0] = [
    #     parent2['variables'][0][i] - crs_flags[i] * rand_num[i] * ratio * (parent2['variables'][0][i] - parent1['variables'][0][i])
    #     for i in range(nvar)]

    # export(parent1['variables'][0], parent2['variables'][0], child1['variables'][0], child2['variables'][0])

    # round and round
    # if any(opt['vartype'] == 2):
    # change = [i for i, v in enumerate(opt['vartype']) if v == 2]
    # for i in change:
    # child1['variables'][i] = round(child1['variables'][i])
    # child2['variables'][i] = round(child2['variables'][i])
    # child1['var'] = child1['var'].round()  # TODO: bunu aktif et
    # child2['var'] = child2['var'].round()

    # todo: varlimit olmalı
    # child1['variables'] = var_limit(child1['variables'], opt['lb'], opt['ub'])
    # child2['variables'] = var_limit(child2['variables'], opt['lb'], opt['ub'])

    # child1['variables_str'] = reshape_variables.reshape_variables(child1['variables'])
    # child2['variables_str'] = reshape_variables.reshape_variables(child2['variables'])

    # if opt['inrepair'] == 1:
    #     child1 = repair_operator.repair_operator(child1, opt)
    #     child2 = repair_operator.repair_operator(child2, opt)

    return child1, child2
