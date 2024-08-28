import random
import numpy as np
import reshape_variables, repair_operator
import pandas as pd
import openpyxl
import copy
from var_limit import var_limit


def export(x, y, z, h):
    x = np.vstack((x, y, z, h))
    wb = openpyxl.load_workbook("crossover_once_sonra.xlsx")
    sheet = wb["Sheet1"]
    # print(sheet.max_row)
    df = pd.DataFrame(x)
    with pd.ExcelWriter("crossover_once_sonra.xlsx", mode="a", engine="openpyxl", if_sheet_exists='overlay') as writer:
        df.to_excel(writer, index=False, header=False, sheet_name="Sheet1", startrow=sheet.max_row)


def crossover_op(pop, opt, state):
    fraction = opt['crossover_fraction']
    # range(0, len(pop), 2)
    if state["currentGen"] < 1000:
        for ind in range(0, len(pop), 2):  # pop_size should be even number
            child1, child2 = crossover_intermediate(pop[ind], pop[ind + 1], fraction, opt)
            pop[ind] = child1
            pop[ind + 1] = child2
    else:
        for ind in range(0, 10):  # pop_size should be even number
            child1, child2 = crossover_intermediate(pop[ind], pop[ind + 1], fraction, opt)
            pop[ind] = child1
            pop[ind + 1] = child2

    return pop


def crossover_intermediate(parent1, parent2, fraction, opt):
    ratio = 1.2000  # matlab'ta baya kapsamlı gibi TODO
    # child1 = parent1.copy()
    # child2 = parent2.copy()
    child1 = copy.deepcopy(parent1)
    child2 = copy.deepcopy(parent2)

    n_var = len(child1['variables'][0])
    crs_flags = np.array([int(random.random() > fraction) for _ in range(n_var)])
    rand_num = np.array([random.random() for _ in range(n_var)])

    # [][0] idi...
    # child1['variables'][0] = [
    #     parent1['variables'][0][i] + crsflag[i] * randnum[i] * ratio * (parent2['variables'][0][i] - parent1['variables'][0][i])
    #     for i in range(nvar)]

    # crs_flags = np.array([1 for _ in range(n_var)])  # hepsi crossovera uğrar
    # test = parent1['variables'][0] + (crs_flags * rand_num * ratio * (parent2['variables'][0] - parent1['variables'][0]))

    child1['variables'] = parent1['variables'] + crs_flags * rand_num * ratio * (parent2['variables'] - parent1['variables'])
    child2['variables'] = parent2['variables'] - crs_flags * rand_num * ratio * (parent2['variables'] - parent1['variables'])

    # child2['variables'][0] = [
    #     parent2['variables'][0][i] - crs_flags[i] * rand_num[i] * ratio * (parent2['variables'][0][i] - parent1['variables'][0][i])
    #     for i in range(nvar)]

    # export(parent1['variables'][0], parent2['variables'][0], child1['variables'][0], child2['variables'][0])

    # round and round
    # if any(opt['vartype'] == 2):
    # change = [i for i, v in enumerate(opt['vartype']) if v == 2]
    # for i in change:
    # child1['variables'][i] = round(child1['variables'][i])
    # child2['variables'][i] = round(child2['variables'][i])
    child1['variables'] = child1['variables'].round()
    child2['variables'] = child2['variables'].round()

    # todo: varlimit olmalı
    child1['variables'] = var_limit(child1['variables'], opt['lb'], opt['ub'])
    child2['variables'] = var_limit(child2['variables'], opt['lb'], opt['ub'])

    child1['variables_str'] = reshape_variables.reshape_variables(child1['variables'])
    child2['variables_str'] = reshape_variables.reshape_variables(child2['variables'])

    # if opt['inrepair'] == 1:
    #     child1 = repair_operator.repair_operator(child1, opt)
    #     child2 = repair_operator.repair_operator(child2, opt)

    return child1, child2
